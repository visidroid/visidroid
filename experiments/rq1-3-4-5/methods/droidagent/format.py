import os 
import pandas as pd
from openpyxl import load_workbook

file = 'droidagent/raw_tasks.xlsx'
wb = load_workbook(file)
sheets = wb.sheetnames

droidagent = pd.read_excel(file, sheet_name='tasks')

all_tasks = pd.DataFrame()

for index, row in droidagent.iterrows():
    hash = row['hash']
    hash = hash[:10]
    nl_actions = ""
        
    if hash in sheets:
        task = pd.read_excel(file, sheet_name=hash)
        actions = task['droidagent']
        actions = actions.dropna()
        
        for index, action in enumerate(actions):
            action = action.replace('Touch', 'touched').replace('Fill', 'filled')
            nl_actions += f'- ACTION {index + 1}: {action}\n'
    
    all_tasks = all_tasks._append({
        "app_name": row['app_name'],
        "task_desc": row['task_desc'],
        "hash": row['hash'],
        "n_actions": len(actions),
        "soa": nl_actions
    }, ignore_index=True)
    
all_tasks = all_tasks.sort_values(by=['app_name', 'task_desc'])
os.chdir('droidagent')
all_tasks.to_excel('all_tasks.xlsx', index=False)


